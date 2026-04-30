"""Inspect SAL_2021_AUST.xlsx - find header row and columns, then list distinct
values of any column whose header looks like GCCSA."""
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "abs" / "SAL_2021_AUST.xlsx"

wb = load_workbook(XLSX, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb[wb.sheetnames[0]]

# Print first 5 rows
rows = ws.iter_rows(min_row=1, max_row=5, values_only=True)
for i, row in enumerate(rows, 1):
    print(f"row{i}:", row[:15])
