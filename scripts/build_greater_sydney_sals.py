"""Determine which SAL_CODE_2021 codes belong to GCCSA = '1GSYD' (Greater Sydney).

Approach:
- Parse MB_2021_AUST.xlsx, build dict: MB_CODE_2021 -> GCCSA_CODE_2021
- Parse SAL_2021_AUST.xlsx, for each SAL_CODE accumulate the GCCSA codes of its mesh blocks
- A SAL is "Greater Sydney" if any of its MBs map to 1GSYD (in practice almost all-or-none)

Outputs:
- abs/greater_sydney_sal_codes.txt  (one SAL_CODE_2021 per line, with SAL_NAME)
"""
import sys
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
ABS = ROOT / "abs"
MB_XLSX = ABS / "MB_2021_AUST.xlsx"
SAL_XLSX = ABS / "SAL_2021_AUST.xlsx"
OUT = ABS / "greater_sydney_sal.tsv"

GCCSA_TARGET = "1GSYD"

def read_headers(ws):
    return [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

print("Loading Mesh Block -> GCCSA map...", file=sys.stderr)
wb = load_workbook(MB_XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
headers = read_headers(ws)
print("MB columns:", headers, file=sys.stderr)
mb_idx = headers.index("MB_CODE_2021")
gccsa_idx = headers.index("GCCSA_CODE_2021")

mb_to_gccsa = {}
n = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    mb_to_gccsa[str(row[mb_idx])] = row[gccsa_idx]
    n += 1
    if n % 100_000 == 0:
        print(f"  {n:,} mb rows", file=sys.stderr)
print(f"Total MB rows: {n:,}", file=sys.stderr)
wb.close()

print("\nWalking SAL allocation file...", file=sys.stderr)
wb = load_workbook(SAL_XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
headers = read_headers(ws)
print("SAL columns:", headers, file=sys.stderr)
mb_i = headers.index("MB_CODE_2021")
sal_i = headers.index("SAL_CODE_2021")
name_i = headers.index("SAL_NAME_2021")

sal_gccsa = defaultdict(Counter)
sal_name = {}
n = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    mb = str(row[mb_i])
    sal = str(row[sal_i])
    sal_name[sal] = row[name_i]
    g = mb_to_gccsa.get(mb)
    if g:
        sal_gccsa[sal][g] += 1
    n += 1
    if n % 200_000 == 0:
        print(f"  {n:,} sal-mb rows", file=sys.stderr)
print(f"Total SAL-MB rows: {n:,}", file=sys.stderr)
wb.close()

# Write out SALs whose dominant GCCSA is target
out_lines = []
gsyd_total = 0
for sal, counter in sal_gccsa.items():
    dominant_g, _ = counter.most_common(1)[0]
    if dominant_g == GCCSA_TARGET:
        out_lines.append(f"{sal}\t{sal_name[sal]}")
        gsyd_total += 1

out_lines.sort(key=lambda s: s.split("\t", 1)[1].lower())
OUT.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
print(f"\nGreater Sydney SALs: {gsyd_total}", file=sys.stderr)
print(f"Wrote {OUT}", file=sys.stderr)
